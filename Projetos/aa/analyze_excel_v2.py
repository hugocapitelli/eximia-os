import openpyxl
import pandas as pd
from pathlib import Path

# Caminho do arquivo
file_path = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"

output_file = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\analise_planilha.txt"

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("=" * 120 + "\n")
    f.write("ANÁLISE DA PLANILHA DE PLANEJAMENTO FINANCEIRO\n")
    f.write("=" * 120 + "\n\n")

    # Carregar o workbook
    wb = openpyxl.load_workbook(file_path, data_only=False)

    f.write(f"📋 ABAS DISPONÍVEIS:\n")
    f.write("-" * 120 + "\n")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        sheet = wb[sheet_name]
        f.write(f"{idx}. {sheet_name} (Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas)\n")

    f.write("\n" + "=" * 120 + "\n")
    f.write("ANÁLISE DETALHADA DE CADA ABA\n")
    f.write("=" * 120 + "\n")

    for sheet_name in wb.sheetnames[:3]:  # Apenas primeiras 3 abas
        sheet = wb[sheet_name]
        f.write(f"\n\n{'='*120}\n")
        f.write(f"ABA: {sheet_name}\n")
        f.write(f"{'='*120}\n\n")
        
        # Tentar usar pandas para visualizar melhor
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=25)
            
            f.write(f"📊 PRIMEIRAS 25 LINHAS:\n")
            f.write("-" * 120 + "\n")
            f.write(df.to_string() + "\n")
        except Exception as e:
            f.write(f"Erro ao ler com pandas: {e}\n")
        
        # Verificar fórmulas
        f.write(f"\n\n🔍 ANÁLISE DE FÓRMULAS (células com fórmulas):\n")
        f.write("-" * 120 + "\n")
        formula_count = 0
        for row_idx, row in enumerate(sheet.iter_rows(max_row=50, max_col=30), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    f.write(f"   Linha {row_idx}, {cell.coordinate}: {cell.value}\n")
                    formula_count += 1
                    if formula_count >= 100:
                        break
            if formula_count >= 100:
                break
        
        if formula_count == 0:
            f.write("   Nenhuma fórmula encontrada nas primeiras 50 linhas.\n")
        else:
            f.write(f"\n   Total de fórmulas encontradas: {formula_count}\n")

    f.write("\n\n" + "=" * 120 + "\n")
    f.write("ANÁLISE CONCLUÍDA\n")
    f.write("=" * 120 + "\n")

print(f"Análise salva em: {output_file}")

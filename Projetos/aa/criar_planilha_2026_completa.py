"""
Script Completo: Criar Planilha 2026 Automaticamente
Duplica TODAS as abas relevantes e ajusta datas para 2026
Remove dados, mantém apenas estrutura e fórmulas
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import shutil

# Configurações
arquivo_original = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"
arquivo_2026 = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2026.xlsx"

# Lista de abas para duplicar e renomear
abas_para_duplicar = {
    "I": "I 2026",
    "DRE": "DRE 2026",
    "BD": "BD 2026",
    "Rec.": "Rec. 2026",
    "D": "D 2026",
    "Op.Fix.": "Op.Fix. 2026",
    "Op.Var.": "Op.Var. 2026",
    "Finan.": "Finan. 2026",
    "D.Adm": "D.Adm 2026",
    "D.Com.": "D.Com. 2026",
    "D.Log.": "D.Log. 2026",
    "Invest.": "Invest. 2026",
    "Divd.": "Divd. 2026",
    "Mov.Cx.": "Mov.Cx. 2026",
    "DFC": "DFC 2026",
    "Proj. DRE": "Proj. DRE 2026",
    "Hist. e Proj.": "Hist. e Proj. 2026"
}

print("=" * 100)
print("🚀 CRIANDO PLANILHA COMPLETA 2026 AUTOMATICAMENTE")
print("=" * 100)

# Copiar arquivo completo
print(f"\n📋 Criando nova planilha: {Path(arquivo_2026).name}")
shutil.copy2(arquivo_original, arquivo_2026)
print("✅ Arquivo base copiado!")

# Carregar planilha
print(f"\n📂 Abrindo arquivo...")
wb = openpyxl.load_workbook(arquivo_2026)

print(f"\n📊 Abas disponíveis: {len(wb.sheetnames)}")

# Processar cada aba
abas_criadas = []
abas_nao_encontradas = []

for aba_origem, aba_destino in abas_para_duplicar.items():
    if aba_origem not in wb.sheetnames:
        abas_nao_encontradas.append(aba_origem)
        continue
    
    print(f"\n🔄 Processando: '{aba_origem}' → '{aba_destino}'")
    
    # Copiar a aba
    sheet_origem = wb[aba_origem]
    sheet_destino = wb.copy_worksheet(sheet_origem)
    sheet_destino.title = aba_destino
    
    # Ajustar datas nas primeiras linhas (geralmente linha 1 e 2)
    datas_ajustadas = 0
    
    for row in [1, 2]:  # Verificar primeiras 2 linhas
        for col in range(1, min(sheet_destino.max_column + 1, 100)):
            cell = sheet_destino.cell(row=row, column=col)
            
            if isinstance(cell.value, datetime):
                try:
                    # Tentar ajustar para 2026
                    if cell.value.year in [2024, 2025]:
                        nova_data = cell.value.replace(year=2026)
                        cell.value = nova_data
                        datas_ajustadas += 1
                except:
                    pass
    
    # Limpar dados (manter fórmulas)
    # Começar da linha 3 para preservar cabeçalhos
    dados_limpos = 0
    
    for row in range(3, min(sheet_destino.max_row + 1, 300)):
        for col in range(1, min(sheet_destino.max_column + 1, 200)):
            cell = sheet_destino.cell(row=row, column=col)
            
            # Se não é fórmula e não é None, limpar
            if cell.data_type != 'f' and cell.value is not None:
                # Verificar se é um label/texto na primeira coluna (geralmente descrições)
                if col == 1 and isinstance(cell.value, str):
                    # Manter labels
                    continue
                
                # Limpar valores
                cell.value = None
                dados_limpos += 1
    
    print(f"   ✅ Aba criada: '{aba_destino}'")
    print(f"   📅 Datas ajustadas: {datas_ajustadas}")
    print(f"   🗑️ Dados limpos: {dados_limpos}")
    
    abas_criadas.append(aba_destino)

# Remover abas antigas (2024/2025) se desejar
print(f"\n🗑️ Deseja remover as abas originais de 2024/2025? (Deixando apenas 2026)")
print("   (Para manter ambas, vamos deixar as originais)")

# Salvar arquivo
print(f"\n💾 Salvando nova planilha 2026...")
wb.save(arquivo_2026)
print(f"✅ Arquivo salvo: {Path(arquivo_2026).name}")

# Resumo
print("\n" + "=" * 100)
print("🎉 PLANILHA 2026 CRIADA COM SUCESSO!")
print("=" * 100)

print(f"\n📌 Resumo:")
print(f"   • Arquivo criado: {Path(arquivo_2026).name}")
print(f"   • Abas criadas: {len(abas_criadas)}")
print(f"   • Dados limpos: mantidas apenas fórmulas e estrutura")

if abas_criadas:
    print(f"\n✅ Abas 2026 criadas:")
    for aba in abas_criadas:
        print(f"   • {aba}")

if abas_nao_encontradas:
    print(f"\n⚠️ Abas não encontradas (puladas):")
    for aba in abas_nao_encontradas:
        print(f"   • {aba}")

print(f"\n📂 Localização: {arquivo_2026}")
print("\n✅ A planilha está pronta para ser preenchida com os dados orçados de 2026!")

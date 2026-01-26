"""
Script Auxiliar: Análise de Fórmulas do Excel
Para ajudar a identificar quais fórmulas precisam ser copiadas para 2026
"""

import openpyxl
import pandas as pd
from pathlib import Path

# Caminho do arquivo
file_path = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"

print("=" * 100)
print("ANÁLISE DE FÓRMULAS PARA REPLICAÇÃO 2026")
print("=" * 100)

# Carregar o workbook SEM data_only para ver as fórmulas
wb = openpyxl.load_workbook(file_path, data_only=False)

# Focar na aba "I" que é a base para 2026
sheet = wb["I"]

print(f"\n📋 Analisando a aba 'I' (Layout 2025)")
print(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")

print("\n" + "=" * 100)
print("FÓRMULAS ENCONTRADAS (Primeiras 100)")
print("=" * 100)

formulas_found = []

# Procurar por fórmulas
for row in sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=15):
    for cell in row:
        try:
            # Verificar se é uma fórmula
            if cell.data_type == 'f':  # 'f' = formula
                formulas_found.append({
                    'Célula': cell.coordinate,
                    'Fórmula': cell.value,
                    'Valor': cell.value if hasattr(cell, 'value') else 'N/A'
                })
        except:
            pass

if formulas_found:
    print(f"\n✅ Encontradas {len(formulas_found)} fórmulas!\n")
    for idx, formula in enumerate(formulas_found[:100], 1):
        print(f"{idx}. Célula {formula['Célula']}: {formula['Fórmula']}")
else:
    print("\n⚠️ Nenhuma fórmula encontrada nas primeiras 50 linhas.")
    print("\nIsso pode significar:")
    print("1. As fórmulas foram convertidas em valores (data_only=True foi usado ao salvar)")
    print("2. A planilha trabalha com valores fixos e não fórmulas")
    print("3. As fórmulas estão em outras células/abas")

print("\n" + "=" * 100)
print("RECOMENDAÇÕES")
print("=" * 100)

if formulas_found:
    print("""
✅ Para copiar essas fórmulas para 2026:

1. MÉTODO AUTOMÁTICO (Mais Fácil):
   - Clique com botão direito na aba "I"
   - Escolha "Mover ou Copiar"
   - Marque "Criar uma cópia"
   - Renomeie para "I 2026"
   - Ajuste as datas no cabeçalho

2. MÉTODO MANUAL (Célula por Célula):
   - Selecione a célula com fórmula
   - Ctrl+C (copiar)
   - Clique na célula de destino
   - Ctrl+V (colar)
   - Excel ajusta automaticamente as referências relativas
""")
else:
    print("""
⚠️ Como não foram encontradas fórmulas (ou foram convertidas em valores):

1. Verifique se a planilha tem fórmulas em outras abas (DRE, BD, etc.)
2. Talvez os cálculos sejam feitos em outras abas e esta apenas exibe valores
3. Considere copiar a estrutura completa e apenas alterar os dados de entrada

IMPORTANTE: Se a aba "I" não tem fórmulas, talvez ela seja alimentada por outra aba.
Recomendo verificar as abas "DRE", "BD" e outras para entender o fluxo de dados.
""")

print("\n" + "=" * 100)
print("ESTRUTURA DE DADOS DA ABA 'I'")
print("=" * 100)

# Mostrar estrutura simplificada
df = pd.read_excel(file_path, sheet_name="I", nrows=30)
print(f"\nPrimeiras 10 linhas da aba 'I':\n")
print(df.head(10).to_string(max_colwidth=30))

print("\n\n" + "=" * 100)
print("✅ ANÁLISE CONCLUÍDA")
print("=" * 100)
print("\nSe precisar de mais detalhes, você pode:")
print("1. Abrir o arquivo Excel manualmente")
print("2. Clicar em uma célula e verificar a barra de fórmulas")
print("3. Procurar pelo símbolo '=' no início da célula")
print("4. Verificar se há referências do tipo '$A$1' (absolutas) ou 'A1' (relativas)")

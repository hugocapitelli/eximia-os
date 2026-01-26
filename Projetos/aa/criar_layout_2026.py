"""
Script Automático: Criar Layout 2026
Duplica a aba "I" (2025) e cria "I 2026" com datas ajustadas
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
import shutil

# Configurações
arquivo_original = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"
nome_aba_origem = "I"
nome_aba_destino = "I 2026"

# Criar backup antes de modificar
print("=" * 80)
print("🔄 CRIANDO LAYOUT 2026 AUTOMATICAMENTE")
print("=" * 80)

# Fazer backup
backup_path = arquivo_original.replace('.xlsx', '_BACKUP.xlsx')
print(f"\n📋 Criando backup em: {backup_path}")
shutil.copy2(arquivo_original, backup_path)
print("✅ Backup criado com sucesso!")

# Carregar planilha
print(f"\n📂 Abrindo arquivo: {Path(arquivo_original).name}")
wb = openpyxl.load_workbook(arquivo_original)

# Verificar se a aba existe
if nome_aba_origem not in wb.sheetnames:
    print(f"❌ ERRO: Aba '{nome_aba_origem}' não encontrada!")
    print(f"Abas disponíveis: {', '.join(wb.sheetnames)}")
    exit(1)

# Verificar se já existe a aba 2026
if nome_aba_destino in wb.sheetnames:
    print(f"\n⚠️ ATENÇÃO: Aba '{nome_aba_destino}' já existe!")
    resposta = input("Deseja sobrescrever? (s/n): ").lower()
    if resposta == 's':
        print(f"🗑️ Removendo aba existente '{nome_aba_destino}'...")
        wb.remove(wb[nome_aba_destino])
    else:
        print("❌ Operação cancelada.")
        exit(0)

# Copiar a aba
print(f"\n📋 Copiando aba '{nome_aba_origem}'...")
aba_origem = wb[nome_aba_origem]
aba_destino = wb.copy_worksheet(aba_origem)
aba_destino.title = nome_aba_destino
print(f"✅ Aba '{nome_aba_destino}' criada!")

# Ajustar as datas no cabeçalho (primeira linha)
print(f"\n📅 Ajustando datas para 2026...")
contador_ajustes = 0

for col in range(1, aba_destino.max_column + 1):
    cell = aba_destino.cell(row=1, column=col)
    
    # Verificar se é uma data
    if isinstance(cell.value, datetime):
        # Adicionar 1 ano
        nova_data = cell.value.replace(year=2026)
        cell.value = nova_data
        contador_ajustes += 1
        print(f"   Coluna {get_column_letter(col)}: {cell.value.strftime('%b/%Y')}")
    
    # Também verificar a linha 2 (pode ter datas em formato diferente)
    cell2 = aba_destino.cell(row=2, column=col)
    if isinstance(cell2.value, datetime):
        nova_data = cell2.value.replace(year=2026)
        cell2.value = nova_data

print(f"✅ {contador_ajustes} datas ajustadas para 2026")

# Salvar arquivo
print(f"\n💾 Salvando arquivo...")
wb.save(arquivo_original)
print(f"✅ Arquivo salvo com sucesso!")

print("\n" + "=" * 80)
print("🎉 LAYOUT 2026 CRIADO COM SUCESSO!")
print("=" * 80)
print(f"\n📌 Resumo:")
print(f"   • Aba criada: '{nome_aba_destino}'")
print(f"   • Backup salvo em: {Path(backup_path).name}")
print(f"   • Datas ajustadas: {contador_ajustes}")
print(f"\n✅ Abra o arquivo Excel e confira a nova aba '{nome_aba_destino}'!")
print("\n⚠️ Se algo deu errado, você pode restaurar o backup.")

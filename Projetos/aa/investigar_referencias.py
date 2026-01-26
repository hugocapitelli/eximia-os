"""
Script: Investigar Referências Cruzadas entre Abas
Verifica se há fórmulas que referenciam outras abas
"""

import openpyxl
import re

file_path = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"

print("=" * 100)
print("INVESTIGAÇÃO DE REFERÊNCIAS CRUZADAS ENTRE ABAS")
print("=" * 100)

wb = openpyxl.load_workbook(file_path, data_only=False)

# Focar na aba "I" e ver se ela referencia outras abas
print("\n🔍 Analisando aba 'I' (base para 2026)...")
sheet_i = wb["I"]

referencias_saindo = []  # Da aba "I" para outras abas
print("\n📤 Referências SAINDO da aba 'I' (para outras abas):")
print("-" * 100)

for row in sheet_i.iter_rows(max_row=100, max_col=20):
    for cell in row:
        if cell.data_type == 'f' and cell.value:
            # Procurar por referências a outras abas (formato: 'NomeAba'!A1 ou NomeAba!A1)
            if "!" in str(cell.value):
                match = re.findall(r"'?([^'!]+)'?!", str(cell.value))
                if match:
                    for aba_ref in match:
                        if aba_ref != "I":  # Não contar auto-referências
                            referencias_saindo.append({
                                'celula': cell.coordinate,
                                'formula': cell.value,
                                'aba_destino': aba_ref
                            })

if referencias_saindo:
    print(f"✅ Encontradas {len(referencias_saindo)} referências para outras abas:\n")
    for ref in referencias_saindo[:20]:  # Mostrar primeiras 20
        print(f"   {ref['celula']}: {ref['formula'][:80]}... → Aba '{ref['aba_destino']}'")
else:
    print("✅ Nenhuma referência para outras abas encontrada")

# Agora verificar se OUTRAS abas referenciam a aba "I"
print("\n" + "=" * 100)
print("📥 Referências ENTRANDO na aba 'I' (de outras abas):")
print("=" * 100)

referencias_entrando = {}

for sheet_name in ['DRE', 'BD', 'Rec.', 'D']:  # Abas mais importantes
    if sheet_name not in wb.sheetnames:
        continue
    
    print(f"\n🔍 Analisando aba '{sheet_name}'...")
    sheet = wb[sheet_name]
    refs = []
    
    for row in sheet.iter_rows(max_row=100, max_col=50):
        for cell in row:
            if cell.data_type == 'f' and cell.value and "'I'!" in str(cell.value):
                refs.append({
                    'celula': cell.coordinate,
                    'formula': cell.value
                })
    
    if refs:
        referencias_entrando[sheet_name] = refs
        print(f"   ⚠️ Encontradas {len(refs)} referências para a aba 'I'")
        for ref in refs[:5]:  # Mostrar primeiras 5
            print(f"      {ref['celula']}: {ref['formula'][:80]}...")
    else:
        print(f"   ✅ Nenhuma referência para a aba 'I'")

# Resumo
print("\n" + "=" * 100)
print("📊 RESUMO E RECOMENDAÇÕES")
print("=" * 100)

if not referencias_saindo and not referencias_entrando:
    print("\n✅ BOA NOTÍCIA!")
    print("   A aba 'I' não tem dependências com outras abas.")
    print("   Pode duplicar tranquilamente sem preocupações!")
else:
    print("\n⚠️ ATENÇÃO - DEPENDÊNCIAS ENCONTRADAS:")
    
    if referencias_saindo:
        print(f"\n   • A aba 'I' referencia outras abas em {len(referencias_saindo)} células")
        print("   → Ao duplicar, essas referências vão continuar apontando para as mesmas abas")
        print("   → Isso pode estar CORRETO se você quer usar os mesmos dados base")
    
    if referencias_entrando:
        print(f"\n   • As seguintes abas referenciam a aba 'I':")
        for aba, refs in referencias_entrando.items():
            print(f"     - {aba}: {len(refs)} referências")
        print("\n   ⚠️ PROBLEMA: Essas abas vão continuar olhando para 'I' (2025)")
        print("   ⚠️ Elas NÃO vão automaticamente usar 'I 2026'")
        
        print("\n   📌 SOLUÇÃO:")
        print("   1. Criar NOVAS abas para 2026 (ex: 'DRE 2026', 'BD 2026')")
        print("   2. Atualizar as referências nessas novas abas de 'I' → 'I 2026'")
        print("   3. Ou usar script de automação para fazer isso")

print("\n" + "=" * 100)

"""
Script: Investigar Referências Cruzadas
Versão simplificada para verificar dependências
"""

import openpyxl

file_path = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"

print("INVESTIGANDO REFERENCIAS CRUZADAS...")

wb = openpyxl.load_workbook(file_path, data_only=False)
sheet_i = wb["I"]

# Contar referências
refs_saindo = 0
refs_entrando = 0

# Verificar aba I
for row in sheet_i.iter_rows(max_row=50):
    for cell in row:
        if cell.data_type == 'f' and cell.value and "!" in str(cell.value):
            refs_saindo += 1

# Verificar DRE
if "DRE" in wb.sheetnames:
    sheet_dre = wb["DRE"]
    for row in sheet_dre.iter_rows(max_row=50):
        for cell in row:
            if cell.data_type == 'f' and cell.value and "'I'!" in str(cell.value):
                refs_entrando += 1

print(f"\nReferencias da aba I para outras: {refs_saindo}")
print(f"Referencas do DRE para I: {refs_entrando}")

if refs_saindo == 0 and refs_entrando == 0:
    print("\nBOA NOTICIA: Sem dependencias criticas!")
else:
    print("\nATENCAO: Existem dependencias entre abas!")
    print("Pode ser necessario ajustar referencias manualmente")

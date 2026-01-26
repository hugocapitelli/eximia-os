import openpyxl
import pandas as pd
from pathlib import Path

# Caminho do arquivo
file_path = r"c:\Users\hugoc\OneDrive\Área de Trabalho\exímIA Ventures\eximIA.OS\Projetos\aa\Planejamento Financeiro_Massas Lott_2024_V20 (version Final) Dezembro 2025.xlsx"

print("=" * 80)
print("ANÁLISE DA PLANILHA DE PLANEJAMENTO FINANCEIRO")
print("=" * 80)

# Carregar o workbook
wb = openpyxl.load_workbook(file_path, data_only=False)

print(f"\n📋 ABAS DISPONÍVEIS:")
print("-" * 80)
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    print(f"{idx}. {sheet_name} (Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas)")

print("\n" + "=" * 80)
print("ANÁLISE DETALHADA DE CADA ABA")
print("=" * 80)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n\n{'='*80}")
    print(f"ABA: {sheet_name}")
    print(f"{'='*80}")
    
    # Tentar usar pandas para visualizar melhor
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    print(f"\n📊 PRIMEIRAS 20 LINHAS:")
    print("-" * 80)
    print(df.head(20).to_string())
    
    # Verificar fórmulas na primeira aba
    if sheet == wb.worksheets[0]:
        print(f"\n🔍 ANÁLISE DE FÓRMULAS (primeiras 50 células com fórmulas):")
        print("-" * 80)
        formula_count = 0
        for row in sheet.iter_rows(max_row=30, max_col=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"   {cell.coordinate}: {cell.value}")
                    formula_count += 1
                    if formula_count >= 50:
                        break
            if formula_count >= 50:
                break
        
        if formula_count == 0:
            print("   Nenhuma fórmula encontrada nas primeiras linhas.")

print("\n\n" + "=" * 80)
print("ANÁLISE CONCLUÍDA")
print("=" * 80)
